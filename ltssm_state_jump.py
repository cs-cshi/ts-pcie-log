from general_log_process import preprocess_log
import openpyxl
import itertools
import os


def extract_ltssm_info(log_path, start_str, end_str):
    """Extract state machine information and store it in Excel"""
    # start_str = "skt_id = 0x0"
    # end_str = "power on..."
    start_flag = False

    total_ltssm_group = []
    current_ltssm_group = []

    total_0x20560 = []
    current_0x20560 = []

    total_ltssm_st_value = []
    current_ltssm_st_value = []

    total_histogram_data = []
    current_histogram_data = []

    total_histogram_data2_height = []
    current_histogram_data2_height = []

    total_eq_obj_value = []
    current_eq_obj_value = []

    total_eq_obj_tap4_value = []
    current_eq_obj_tap4_value = []

    total_eq_obj_tap5_value = []
    current_eq_obj_tap5_value = []

    total_eq_obj_tap6_value = []
    current_eq_obj_tap6_value = []

    total_eq_obj_ffe_taps_value = []
    current_eq_obj_ffe_taps_value = []

    total_eye_height_value = []
    current_eye_height_value = []

    total_adc_os_cal_dig_cor_gen_value = []
    current_adc_os_cal_dig_cor_gen_value = []

    total_adc_os_cal_actl_gen_value = []
    current_adc_os_cal_actl_gen_value = []

    total_take_it_out_value = []
    current_take_it_out_value = []

    total_take_it_out_early_value = []
    current_take_it_out_early_value = []

    # module_title = ["ltssm state track"]

    run_times = 0

    # 打开或创建 Excel 工作簿
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'PCIe LTSSM'

    # 读取日志文件内容
    with open(log_path, 'r', errors='ignore') as file:
        log_lines = file.readlines()

        # 逐行处理
        for line in log_lines:
            if start_str in line:
                run_times += 1
                start_flag = True

            # if "******************" in line:
            #     last_star_index = line.rfind('*')
            #     last_star_index += 1
            #     line = line[last_star_index:]
            #     module_group.append(line)

            if start_flag:
                line = line.strip()
                if "ltssm state track is:" in line:
                    current_ltssm_group.append(line.split(":")[-1])
                elif "0x20560" in line:
                    current_0x20560.append(line)
                elif "ltssm_st value" in line:
                    current_ltssm_st_value.append(line)
                elif "histogram_data value" in line:
                    current_histogram_data.append(line)
                elif "histogram_data2 value" in line:
                    current_histogram_data2_height.append(line)
                elif "eq obj value" in line:
                    current_eq_obj_value.append(line)
                elif "eq obj tap4 value" in line:
                    current_eq_obj_tap4_value.append(line)
                elif "eq obj tap5 value" in line:
                    current_eq_obj_tap5_value.append(line)
                elif "eq obj tap6 value" in line:
                    current_eq_obj_tap6_value.append(line)
                elif "eq obj ffe taps value" in line:
                    current_eq_obj_ffe_taps_value.append(line)
                elif "eye height value" in line:
                    current_eye_height_value.append(line)
                elif "adc os cal dig cor gen value" in line:
                    current_adc_os_cal_dig_cor_gen_value.append(line)
                elif "adc os cal actl gen value" in line:
                    current_adc_os_cal_actl_gen_value.append(line)
                elif "take it out value" in line:
                    current_take_it_out_value.append(line)
                elif "take it out early value" in line:
                    current_take_it_out_early_value.append(line)

            if end_str in line:
                if current_ltssm_group:
                    total_ltssm_group.append(current_ltssm_group.copy())

                if current_0x20560:
                    total_0x20560.append(current_0x20560.copy())

                if current_ltssm_st_value:
                    total_ltssm_st_value.append(current_ltssm_st_value.copy())

                if current_histogram_data:
                    total_histogram_data.append(current_histogram_data.copy())

                if current_histogram_data2_height:
                    total_histogram_data2_height.append(current_histogram_data2_height.copy())

                if current_eq_obj_value:
                    total_eq_obj_value.append(current_eq_obj_value.copy())

                if current_eq_obj_tap4_value:
                    total_eq_obj_tap4_value.append(current_eq_obj_tap4_value.copy())

                if current_eq_obj_tap5_value:
                    total_eq_obj_tap5_value.append(current_eq_obj_tap5_value.copy())

                if current_eq_obj_tap6_value:
                    total_eq_obj_tap6_value.append(current_eq_obj_tap6_value.copy())

                if current_eq_obj_ffe_taps_value:
                    total_eq_obj_ffe_taps_value.append(current_eq_obj_ffe_taps_value.copy())

                if current_eye_height_value:
                    total_eye_height_value.append(current_eye_height_value.copy())

                if current_adc_os_cal_dig_cor_gen_value:
                    total_adc_os_cal_dig_cor_gen_value.append(current_adc_os_cal_dig_cor_gen_value.copy())

                if current_adc_os_cal_actl_gen_value:
                    total_adc_os_cal_actl_gen_value.append(current_adc_os_cal_actl_gen_value.copy())

                if current_take_it_out_value:
                    total_take_it_out_value.append(current_take_it_out_value.copy())

                if current_take_it_out_early_value:
                    total_take_it_out_early_value.append(current_take_it_out_early_value.copy())

                current_ltssm_group.clear()
                current_0x20560.clear()
                current_ltssm_st_value.clear()
                current_histogram_data.clear()
                current_histogram_data2_height.clear()
                current_eq_obj_value.clear()
                current_eq_obj_tap4_value.clear()
                current_eq_obj_tap5_value.clear()
                current_eq_obj_tap6_value.clear()
                current_eq_obj_ffe_taps_value.clear()
                current_eye_height_value.clear()
                current_adc_os_cal_dig_cor_gen_value.clear()
                current_adc_os_cal_actl_gen_value.clear()
                current_take_it_out_value.clear()
                current_take_it_out_early_value.clear()

                start_flag = False

    # 将数据存储到 Excel 中
    excel_row = 1
    excel_column = 1

    sheet.cell(row=excel_row, column=excel_column, value="recovery times")
    excel_row += 1

    sheet.cell(row=excel_row, column=excel_column, value="data2 eye height:")

    for sub_ltssm_group, sub_0x20560, sub_ltssm_st_value, sub_histogram_data, sub_histogram_data2_height, sub_eq_obj_value, sub_eq_obj_tap4_value, sub_eq_obj_tap5_value, sub_eq_obj_tap6_value, sub_eq_obj_ffe_taps_value, sub_eye_height_value, sub_adc_os_cal_dig_cor_gen_value, sub_adc_os_cal_actl_gen_value, sub_take_it_out_value, sub_take_it_out_early_value in itertools.zip_longest(
            total_ltssm_group, total_0x20560, total_ltssm_st_value,
            total_histogram_data, total_histogram_data2_height,
            total_eq_obj_value, total_eq_obj_tap4_value,
            total_eq_obj_tap5_value, total_eq_obj_tap6_value,
            total_eq_obj_ffe_taps_value, total_eye_height_value,
            total_adc_os_cal_dig_cor_gen_value,
            total_adc_os_cal_actl_gen_value, total_take_it_out_value,
            total_take_it_out_early_value):

        excel_column += 1
        excel_row = 1

        recovery_state_times = determine_ltssm_stable(sub_ltssm_group)
        sheet.cell(row=excel_row, column=excel_column, value=str(recovery_state_times))
        # print(recovery_state_times)
        excel_row += 1

        eye_height = get_eye_height(sub_histogram_data2_height)

        for i, eye_value in enumerate(eye_height):
            # print(eye_value)
            sheet.cell(row=excel_row, column=excel_column, value=eye_value)
            excel_row += 1

        for i in range(3):
            sheet.cell(row=excel_row, column=excel_column, value=" ")
            excel_row += 1

        if sub_ltssm_group:
            for item in sub_ltssm_group:
                sheet.cell(row=excel_row, column=excel_column, value=item)
                excel_row += 1

            for i in range(3):
                sheet.cell(row=excel_row, column=excel_column, value=" ")
                excel_row += 1

        if sub_0x20560:
            for item in sub_0x20560:
                sheet.cell(row=excel_row, column=excel_column, value=item)
                excel_row += 1

            for i in range(3):
                sheet.cell(row=excel_row, column=excel_column, value=" ")
                excel_row += 1

        if sub_ltssm_st_value:
            for item in sub_ltssm_st_value:
                sheet.cell(row=excel_row, column=excel_column, value=item)
                excel_row += 1

            for i in range(3):
                sheet.cell(row=excel_row, column=excel_column, value=" ")
                excel_row += 1

        if sub_histogram_data:
            for item in sub_histogram_data:
                sheet.cell(row=excel_row, column=excel_column, value=item)
                excel_row += 1

            for i in range(3):
                sheet.cell(row=excel_row, column=excel_column, value=" ")
                excel_row += 1

        if sub_histogram_data2_height:
            for item in sub_histogram_data2_height:
                sheet.cell(row=excel_row, column=excel_column, value=item)
                excel_row += 1

            for i in range(3):
                sheet.cell(row=excel_row, column=excel_column, value=" ")
                excel_row += 1

        if sub_eq_obj_value:
            for item in sub_eq_obj_value:
                sheet.cell(row=excel_row, column=excel_column, value=item)
                excel_row += 1

            for i in range(3):
                sheet.cell(row=excel_row, column=excel_column, value=" ")
                excel_row += 1

        if sub_eq_obj_tap4_value:
            for item in sub_eq_obj_tap4_value:
                sheet.cell(row=excel_row, column=excel_column, value=item)
                excel_row += 1

            for i in range(3):
                sheet.cell(row=excel_row, column=excel_column, value=" ")
                excel_row += 1

        if sub_eq_obj_tap5_value:
            for item in sub_eq_obj_tap5_value:
                sheet.cell(row=excel_row, column=excel_column, value=item)
                excel_row += 1

            for i in range(3):
                sheet.cell(row=excel_row, column=excel_column, value=" ")
                excel_row += 1

        if sub_eq_obj_tap6_value:
            for item in sub_eq_obj_tap6_value:
                sheet.cell(row=excel_row, column=excel_column, value=item)
                excel_row += 1

            for i in range(3):
                sheet.cell(row=excel_row, column=excel_column, value=" ")
                excel_row += 1

        if sub_eq_obj_ffe_taps_value:
            for item in sub_eq_obj_ffe_taps_value:
                sheet.cell(row=excel_row, column=excel_column, value=item)
                excel_row += 1

            for i in range(3):
                sheet.cell(row=excel_row, column=excel_column, value=" ")
                excel_row += 1

        if sub_eye_height_value:
            for item in sub_eye_height_value:
                sheet.cell(row=excel_row, column=excel_column, value=item)
                excel_row += 1

            for i in range(3):
                sheet.cell(row=excel_row, column=excel_column, value=" ")
                excel_row += 1

        if sub_adc_os_cal_dig_cor_gen_value:
            for item in sub_adc_os_cal_dig_cor_gen_value:
                sheet.cell(row=excel_row, column=excel_column, value=item)
                excel_row += 1

            for i in range(3):
                sheet.cell(row=excel_row, column=excel_column, value=" ")
                excel_row += 1

        if sub_adc_os_cal_actl_gen_value:
            for item in sub_adc_os_cal_actl_gen_value:
                sheet.cell(row=excel_row, column=excel_column, value=item)
                excel_row += 1

            for i in range(3):
                sheet.cell(row=excel_row, column=excel_column, value=" ")
                excel_row += 1

        if sub_take_it_out_value:
            for item in sub_take_it_out_value:
                sheet.cell(row=excel_row, column=excel_column, value=item)
                excel_row += 1

            for i in range(3):
                sheet.cell(row=excel_row, column=excel_column, value=" ")
                excel_row += 1

        if sub_take_it_out_early_value:
            for item in sub_take_it_out_early_value:
                sheet.cell(row=excel_row, column=excel_column, value=item)
                excel_row += 1

            for i in range(3):
                sheet.cell(row=excel_row, column=excel_column, value=" ")
                excel_row += 1

        if sub_eye_height_phase23_value:
            for item in sub_eye_height_phase23_value:
                sheet.cell(row=excel_row, column=excel_column, value=item)
                excel_row += 1

            for i in range(3):
                sheet.cell(row=excel_row, column=excel_column, value=" ")
                excel_row += 1

        if sub_eye_height_l0_value:
            for item in sub_eye_height_l0_value:
                sheet.cell(row=excel_row, column=excel_column, value=item)
                excel_row += 1

    # for i, ltssm_group in enumerate(total_ltssm_group):
    #     sheet.cell(row=excel_row, column=excel_column, value="   ")
    #     for j, group in enumerate(ltssm_group):
    #         sheet.cell(row=excel_row, column=excel_column, value=group)
    wb.save(log_path.split(".log")[0] + ".xlsx")
    # wb.save(r'E:\files\phytium\TianshuS\log\nicX1-10times.log.xlsx')


def get_eye_height(eye_height_list):
    """计算眼图睁开的高度，连续 0x0 个数"""

    if not eye_height_list:
        return []

    max_count_list = []

    eye_height_list_split = split_eye_height_by_lane(eye_height_list)

    for sub_list in eye_height_list_split:
        current_count = 0
        max_count = 0
        for current_eye_value in sub_list:
            # current_eye_value = item.split(":")

            if current_eye_value == "0x0":
                current_count += 1
            else:
                current_count = 0

            if current_count > max_count:
                max_count = current_count

        max_count_list.append(max_count)

    return max_count_list


def split_eye_height_by_lane(eye_height_list):
    lanes_value = {}
    for item in eye_height_list:
        lane, value = item.split(":")
        lane.strip()
        value.strip()
        if lane in lanes_value:
            lanes_value[lane].append(value)
        else:
            lanes_value[lane] = [value]

    return list(lanes_value.values())


def determine_ltssm_stable(ltssm_trace):
    """
    Determine whether the ltssm state machine is normal
    :param ltssm_trace:
    :return:
    """

    if not ltssm_trace:
        return 0

    recovery_state_times = 0
    for i, group in enumerate(ltssm_trace):
        group = group.strip()
        if i > 10 and group != "0x0":
            # print(group)
            recovery_state_times += 1

    return recovery_state_times


def fill_cell(sheet, cell_row, cell_column, cell_value, nums):
    """
    :param sheet:
    :param cell_row:
    :param cell_column:
    :param cell_value:
    :param nums:
    :return:
    """
    for i in range(nums):
        sheet.cell(row=cell_row, column=cell_column, value=cell_value)


def test():
    begin_str = "skt_id = 0x0"
    end_str = "power(1:poweroff,2:reset):"
    # file_path = preprocess_log(r"E:\files\phytium\TianshuS\log\nicX1-10times.log", begin_str,                               end_str)
    # print(file_path)
    # extract_ltssm_info(file_path, begin_str, end_str)

    # 指定目录
    # directory_path = r'E:\files\phytium\TianshuS\log\20240829\nic-er_90v'
    directory_path = r'D:\zygfile\desktoop\S5000c-80\问题排查及数据\reg-print\0830-fw12-log'
    # 获取目录中的所有文件
    files = [os.path.join(directory_path, file) for file in os.listdir(directory_path) if
             os.path.isfile(os.path.join(directory_path, file))]

    # 遍历文件并执行处理函数
    for file_path in files:
        if file_path.endswith(".log"):
            # print(file_path)
            # preprocessed_log_path = os.path.splitext(file_path)[0] + '_preprocessed.log'
            # preprocess_log(file_path, preprocessed_log_path, start_str, end_str)
            # record_error_info_b2b(file_path)
            result_path = preprocess_log(file_path, begin_str, end_str)
            extract_ltssm_info(result_path, begin_str, end_str)
            # print(result_path)


test()
